from os import remove
from os.path import abspath
from json import dumps
import openpyxl
from wget import download


download(
    'https://files.x1z53.ru/import_substitution/import_substitution.xlsx',
    'import_substitution.xlsx'
)

# wookbook = openpyxl.load_workbook(filename)
wookbook = openpyxl.load_workbook('import_substitution.xlsx')
worksheet = wookbook.active
result = {}

for i in range(1, worksheet.max_row):
    name, substitution, *classes = [cell[i].value for cell in worksheet.iter_cols(
        1, worksheet.max_column) if cell[i].value]
    subsititution = substitution.replace('"', "'")
    classes = [i.replace('"', "'") for i in classes]
    if name in result:
        result[name]['substitution'] = '|'.join(
            set([*result[name]['substitution'].split('|'), substitution]))
        result[name]['classes'] = '|'.join(
            set([*result[name]['classes'].split('|'), *classes]))
    else:
        result[name] = {}
        result[name]['substitution'] = substitution
        result[name]['classes'] = '|'.join(classes)

result = dumps(result)
local_file = '/'.join(abspath(__file__).split('/')
                      [:-1]) + '/../pages/import_substitution.json'
open(local_file, 'w').write(result)
remove('import_substitution.xlsx')
