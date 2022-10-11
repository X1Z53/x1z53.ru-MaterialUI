from time import ctime, strptime, strftime, localtime
from string import ascii_uppercase
import openpyxl


def main(database, content, *_):
    print('> Start updating:', strftime('%H:%M:%S', localtime()))

    print('> Reading a book')
    sheet = list(openpyxl.load_workbook('book.xlsx'))[0]
    alphabet = ascii_uppercase[2:sheet.max_column]

    full_list = {}
    max_row = sheet.max_row
    for row in range(1, max_row):
        title = sheet['A'][row].value
        try:
            full_list[title]
        except KeyError:
            full_list[title] = [set(), set()]

        full_list[title][0].add(sheet['B'][row].value)

        for col in alphabet:
            if sheet[col][row].value:
                full_list[title][1].add(sheet[col][row].value)

    for key, values in full_list.items():
        for i in range(len(values)):
            full_list[key][i] = ' | '.join(values[i])

    print('> Filling in the table')
    table = database.cursor()
    table.execute('SELECT * FROM import_substitution;')
    table.execute('DELETE FROM import_substitution;')

    for name, parametrs in full_list.items():
        request = f"""INSERT INTO import_substitution({','.join(content['parametrs'])}) VALUES({f"'{name}'" if "'" not in name else f'"{name}"'}, {','.join([f"'{parametr}'" if "'" not in parametr else f'"{parametr}"' for parametr in parametrs])});"""
        table.execute(request)

    database.commit()
    print('> Filling complete')

    print('> Updating has ended:', strftime('%H:%M:%S', localtime()))
